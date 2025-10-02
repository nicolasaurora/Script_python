import requests
import os
from dotenv import load_dotenv
from datetime import datetime, timedelta
from requests.auth import HTTPBasicAuth
import openpyxl
import json
import time

# Cargar credenciales
load_dotenv()
CLIENT_ID = os.getenv("CLIENT_ID")
SECRET_ID = os.getenv("CLIENT_SECRET")

# Endpoint de autenticación
TOKEN_URL = "https://xubio.com/API/1.1/TokenEndpoint"
BASE_URL = "https://xubio.com/API/1.1"

# ENDPOINTS
ENDPOINTS_FUNCIONALES = {
    "asiento_contable" : "asientoContableManualBean",
    "clientes": "clienteBean",
    "retenciones": "retencionBean", 
    "cuentas": "cuenta",
    "factura_compra": "comprobanteCompraBean",
    "factura_venta": "comprobanteVentaBean",
    "cobros" : "cobranzaBean",
    "pagos": "pagoBean",  
    "categorias_cuentas": "categoriaCuenta"  
}


CONFIG_FECHAS = {
    "FECHA_DESDE": "2024-01-01",  
    "FECHA_HASTA": None,  
}

def get_token():
    """Genera y devuelve un access_token válido"""
    print("🔑 Generando token de acceso...")
    data = {"grant_type": "client_credentials",
            "scope":"api_auth"}

    try:
        response = requests.post(
            TOKEN_URL,
            data=data,
            auth=HTTPBasicAuth(CLIENT_ID, SECRET_ID),
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=30
        )
        response.raise_for_status()
        token_data = response.json()
        return token_data["access_token"]
    except requests.exceptions.RequestException as e:
        print(f"❌ Error obteniendo token: {e}")
        raise

def get_data_monthly_chunks_only(token, endpoint_name, endpoint, fecha_desde=None):
    """MÉTODO ÚNICO: Obtiene datos dividiendo por meses (más efectivo y rápido)"""
    if fecha_desde is None:
        fecha_desde = CONFIG_FECHAS["FECHA_DESDE"]
        
    print(f"📅 Descargando {endpoint_name} mes por mes desde {fecha_desde}")
    
    all_data = []
    start_date = datetime.strptime(fecha_desde, "%Y-%m-%d")
    end_date = datetime.now()
    
    current_date = start_date
    
    while current_date <= end_date:
        # Calcular fin del mes
        if current_date.month == 12:
            next_month = current_date.replace(year=current_date.year + 1, month=1)
        else:
            next_month = current_date.replace(month=current_date.month + 1)
        
        month_end = next_month - timedelta(days=1)
        if month_end > end_date:
            month_end = end_date
        
        month_start_str = current_date.strftime("%Y-%m-%d")
        month_end_str = month_end.strftime("%Y-%m-%d")
        
        print(f"   📊 {current_date.strftime('%B %Y')}: {month_start_str} → {month_end_str}", end=" ")
        
        url = f"{BASE_URL}/{endpoint}"
        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/json"
        }
        
        params = {
            "fechaDesde": month_start_str,
            "fechaHasta": month_end_str
        }
        
        try:
            response = requests.get(url, headers=headers, params=params, timeout=90)
            
            if response.status_code == 200:
                try:
                    data = response.json()
                    monthly_data = data if isinstance(data, list) else [data] if data else []
                    
                    if monthly_data:
                        all_data.extend(monthly_data)
                        print(f"→ ✅ {len(monthly_data)} registros")
                    else:
                        print(f"→ ⚪ Sin datos")
                        
                except json.JSONDecodeError:
                    print(f"→ ⚠️ JSON inválido")
            else:
                print(f"→ ❌ Error {response.status_code}")
                
        except requests.exceptions.Timeout:
            print(f"→ ⏰ Timeout")
        except Exception as e:
            print(f"→ ⚠️ Error: {e}")
        
        
        current_date = next_month
        time.sleep(0.5)  
    
    print(f"   🎯 Total {endpoint_name}: {len(all_data)} registros\n")
    return all_data

def get_data_simple_for_catalogs(token, endpoint_name, endpoint):
    """Método simple para endpoints que no usan fechas (como clientes, cuentas)"""
    print(f"📋 Descargando {endpoint_name} (catálogo sin fechas)")
    
    url = f"{BASE_URL}/{endpoint}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=180)
        
        if response.status_code == 200:
            try:
                data = response.json()
                result = data if isinstance(data, list) else [data] if data else []
                print(f"   ✅ {len(result)} registros obtenidos\n")
                return result
            except json.JSONDecodeError:
                print(f"   ⚠️ JSON inválido\n")
                return []
        else:
            print(f"   ❌ Error {response.status_code}\n")
            return []
            
    except Exception as e:
        print(f"   ❌ Error: {e}\n")
        return []

def get_asientos_contables_con_detalle_mejorado(token, endpoint_name, endpoint):
    """Método especial para asientos contables con diagnóstico automático del campo ID"""
    print(f"📊 Descargando {endpoint_name} con detalle completo (DIAGNÓSTICO)")
    
    url = f"{BASE_URL}/{endpoint}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }
    
    
    print(f"   🔄 Paso 1: Obteniendo cabeceras de asientos", end=" ")
    
    try:
        response = requests.get(url, headers=headers, timeout=120)
        
        if response.status_code != 200:
            print(f"→ ❌ Error {response.status_code}")
            return []
            
        try:
            data = response.json()
            cabeceras = data if isinstance(data, list) else [data] if data else []
            print(f"→ ✅ {len(cabeceras)} cabeceras obtenidas")
            
            if not cabeceras:
                print(f"   ⚪ No hay asientos para procesar\n")
                return []
                
        except json.JSONDecodeError:
            print(f"→ ⚠️ JSON inválido")
            return []
            
    except Exception as e:
        print(f"→ ❌ Error: {e}")
        return []
    
    # DIAGNÓSTICO: Analizar estructura de los primeros asientos
    print(f"   🔍 DIAGNÓSTICO: Analizando estructura de los asientos...")
    
    campos_por_frecuencia = {}
    posibles_ids = []
    
    # Analizar los primeros 5 asientos para identificar patrones
    muestra = cabeceras[:5]
    
    for i, asiento in enumerate(muestra):
        if isinstance(asiento, dict):
            print(f"      📋 Asiento {i+1} - Campos disponibles:")
            for campo, valor in asiento.items():
                
                if campo not in campos_por_frecuencia:
                    campos_por_frecuencia[campo] = 0
                campos_por_frecuencia[campo] += 1
                
                
                valor_str = str(valor)[:50] + "..." if len(str(valor)) > 50 else str(valor)
                print(f"         • {campo}: {valor_str}")
                
                
                if ('id' in campo.lower() or 
                    'numero' in campo.lower() or 
                    'codigo' in campo.lower() or
                    'transaccion' in campo.lower()):
                    if campo not in posibles_ids:
                        posibles_ids.append(campo)
            print()
    
    print(f"   🎯 CAMPOS IDENTIFICADOS COMO POSIBLES IDs:")
    if posibles_ids:
        for campo in posibles_ids:
            frecuencia = campos_por_frecuencia.get(campo, 0)
            print(f"      • {campo} (presente en {frecuencia}/{len(muestra)} asientos)")
    else:
        print(f"      ⚠️  No se encontraron campos ID obvios")
        print(f"      📝 Campos más comunes:")
        campos_ordenados = sorted(campos_por_frecuencia.items(), 
                                key=lambda x: x[1], 
                                reverse=True)[:10]
        for campo, freq in campos_ordenados:
            print(f"         • {campo}: {freq}/{len(muestra)}")
    
   
    id_field = None
    
   
    candidatos_id = [
        'transaccionId', 'id', 'idAsiento', 'numeroAsiento', 
        'asientoId', 'codigoAsiento', 'numeroTransaccion',
        'transactionId', 'entryId', 'journalId', 'accountingEntryId',
        'codigo', 'numero', 'clave', 'key'
    ]
    
    
    for candidato in candidatos_id:
        if candidato in posibles_ids:
            id_field = candidato
            break
    
    
    if not id_field:
        for campo in posibles_ids:
            if 'id' in campo.lower():
                id_field = campo
                break
    
    if not id_field and posibles_ids:
        id_field = posibles_ids[0]
    
    if id_field:
        print(f"   ✅ CAMPO ID SELECCIONADO: '{id_field}'")
    else:
        print(f"   ⚠️  NO SE PUDO IDENTIFICAR CAMPO ID")
        print(f"      💡 Continuando solo con cabeceras...")
        print(f"   🎯 Total {endpoint_name}: {len(cabeceras)} registros (solo cabeceras)\n")
        return cabeceras
    
    # Paso 2: Obtener el detalle de cada asiento usando el campo ID identificado
    print(f"   🔄 Paso 2: Obteniendo detalle usando campo '{id_field}'...")
    asientos_completos = []
    exitosos = 0
    errores = 0
    
    for i, asiento in enumerate(cabeceras, 1):
        
        transaccion_id = asiento.get(id_field)
        
        if not transaccion_id:
            asientos_completos.append(asiento)
            errores += 1
            continue
        
        
        detalle_url = f"{BASE_URL}/{endpoint}/{transaccion_id}"
        
        try:
            detalle_response = requests.get(detalle_url, headers=headers, timeout=30)
            
            if detalle_response.status_code == 200:
                try:
                    detalle = detalle_response.json()
                    
                    asientos_completos.append(detalle)
                    exitosos += 1
                    
                    if i % 50 == 0:  
                        print(f"      [{i}/{len(cabeceras)}] ✅ {exitosos} exitosos, {errores} errores")
                        
                except json.JSONDecodeError:
                    asientos_completos.append(asiento)
                    errores += 1
            else:
                
                if errores < 3:
                    print(f"      [{i}/{len(cabeceras)}] ❌ Error {detalle_response.status_code} para ID '{transaccion_id}'")
                asientos_completos.append(asiento)
                errores += 1
                
        except requests.exceptions.Timeout:
            asientos_completos.append(asiento)
            errores += 1
        except Exception as e:
            if errores < 3:  
                print(f"      [{i}/{len(cabeceras)}] ⚠️  Error para ID '{transaccion_id}': {str(e)[:50]}...")
            asientos_completos.append(asiento)
            errores += 1
        
        
        if i % 10 == 0:
            time.sleep(0.5)
    
    print(f"   ✅ Procesamiento completado:")
    print(f"      • Total asientos: {len(asientos_completos)}")
    print(f"      • Con detalle completo: {exitosos}")
    print(f"      • Solo cabeceras: {errores}")
    print(f"      • Campo ID usado: '{id_field}'")
    print(f"   🎯 Total {endpoint_name}: {len(asientos_completos)} registros\n")
    
    return asientos_completos

def get_asientos_contables_debug_solo(token, endpoint_name, endpoint):
    """Versión de debugging que SOLO muestra la estructura sin procesar detalles"""
    print(f"🐛 DEBUG: Analizando estructura de {endpoint_name}")
    
    url = f"{BASE_URL}/{endpoint}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=30)
        
        if response.status_code != 200:
            print(f"❌ Error {response.status_code}")
            return []
            
        data = response.json()
        cabeceras = data if isinstance(data, list) else [data] if data else []
        
        if not cabeceras:
            print(f"⚪ No hay datos")
            return []
        
        print(f"✅ {len(cabeceras)} registros obtenidos")
        print(f"\n📋 ESTRUCTURA DEL PRIMER ASIENTO:")
        print("-" * 50)
        
        primer_asiento = cabeceras[0]
        if isinstance(primer_asiento, dict):
            for campo, valor in primer_asiento.items():
                tipo = type(valor).__name__
                valor_preview = str(valor)[:100] + "..." if len(str(valor)) > 100 else str(valor)
                print(f"• {campo:<25} ({tipo:<10}): {valor_preview}")
        
        
        print(f"\n📄 JSON COMPLETO DEL PRIMER ASIENTO:")
        print("-" * 50)
        print(json.dumps(primer_asiento, indent=2, ensure_ascii=False)[:2000] + "...")
        
        return cabeceras[:10]  
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return []

def aplanar_item_final(item):
    """Convierte cualquier diccionario, lista u objeto anidado a string."""
    if isinstance(item, dict):
        return {k: aplanar_item_final(v) for k, v in item.items()}
    elif isinstance(item, list):
        return json.dumps([aplanar_item_final(x) for x in item], ensure_ascii=False)
    elif isinstance(item, (int, float, str, bool)) or item is None:
        return item
    else:
        return str(item)

def exportar_a_excel_simple(datos_por_recurso, filename="xubio_mensual.xlsx"):
    """Exporta datos con análisis simple de fechas"""
    print(f"💾 Exportando a {filename}...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Hoja de resumen
    summary_ws = wb.create_sheet(title="Resumen")
    summary_ws.append(["Endpoint", "Registros", "Fecha Más Antigua", "Fecha Más Reciente", "Meses Cubiertos"])
    
    for nombre, data in datos_por_recurso.items():
        sheet_name = nombre.replace('_', ' ').title()[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        if not data:
            ws.append(["Sin datos disponibles"])
            summary_ws.append([nombre, 0, "", "", 0])
            continue

        # Analizar fechas
        fechas_encontradas = []
        for item in data:
            if isinstance(item, dict):
                for key, value in item.items():
                    if 'fecha' in key.lower() and isinstance(value, str):
                        try:
                            fecha = datetime.strptime(value[:10], "%Y-%m-%d")
                            fechas_encontradas.append(fecha)
                        except:
                            pass
        
        fecha_min = min(fechas_encontradas).strftime("%Y-%m-%d") if fechas_encontradas else ""
        fecha_max = max(fechas_encontradas).strftime("%Y-%m-%d") if fechas_encontradas else ""
        
        # Calcular meses cubiertos
        meses_cubiertos = 0
        if fechas_encontradas:
            meses_set = set()
            for fecha in fechas_encontradas:
                meses_set.add(fecha.strftime("%Y-%m"))
            meses_cubiertos = len(meses_set)
        
        # Datos
        data_aplanada = [aplanar_item_final(item) for item in data]
        headers = list(data_aplanada[0].keys())
        ws.append(headers)
        
        for item in data_aplanada:
            fila = []
            for col in headers:
                val = item.get(col, "")
                if isinstance(val, (dict, list)):
                    val = json.dumps(val, ensure_ascii=False)
                fila.append(val)
            ws.append(fila)
        
        # Auto-ajustar columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Agregar a resumen
        summary_ws.append([nombre, len(data), fecha_min, fecha_max, meses_cubiertos])

    # Auto-ajustar resumen
    for col in summary_ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        summary_ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)
    print(f"✅ Exportado: {filename}")
    return filename

def generar_reporte_mensual(datos, filename="reporte_mensual.txt"):
    """Genera reporte enfocado en cobertura mensual"""
    with open(filename, "w", encoding="utf-8") as f:
        f.write("REPORTE MENSUAL - XUBIO API\n")
        f.write("=" * 50 + "\n\n")
        f.write(f"Fecha de extracción: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Período: {CONFIG_FECHAS['FECHA_DESDE']} → {datetime.now().strftime('%Y-%m-%d')}\n")
        f.write(f"Método: Chunks mensuales + diagnóstico asientos\n\n")
        
        f.write("RESULTADOS POR ENDPOINT:\n")
        f.write("-" * 30 + "\n")
        
        for nombre, data in datos.items():
            count = len(data) if data else 0
            f.write(f"\n{nombre.upper()}: {count} registros\n")
            
            if data:
                # Analizar cobertura mensual
                meses_con_datos = set()
                fechas = []
                
                for item in data:
                    if isinstance(item, dict):
                        for key, value in item.items():
                            if 'fecha' in key.lower() and isinstance(value, str):
                                try:
                                    fecha = datetime.strptime(value[:10], "%Y-%m-%d")
                                    fechas.append(fecha)
                                    meses_con_datos.add(fecha.strftime("%Y-%m"))
                                except:
                                    pass
                
                if fechas:
                    fecha_min = min(fechas)
                    fecha_max = max(fechas)
                    f.write(f"  Rango: {fecha_min.strftime('%Y-%m-%d')} → {fecha_max.strftime('%Y-%m-%d')}\n")
                    f.write(f"  Meses con datos: {len(meses_con_datos)}\n")
                    
                    # Listar meses
                    meses_ordenados = sorted(list(meses_con_datos))
                    f.write(f"  Meses: {', '.join(meses_ordenados)}\n")
                else:
                    f.write(f"  Sin fechas detectadas en los datos\n")
            else:
                f.write(f"  Sin datos obtenidos\n")

    print(f"📄 Reporte guardado: {filename}")

def main():
    """Función principal con diagnóstico para asientos contables"""
    try:
        print("🚀 XUBIO API - DESCARGA CON DIAGNÓSTICO DE ASIENTOS")
        print("=" * 60)
        print("📅 MÉTODO: Chunks mensuales + diagnóstico automático")
        print(f"🎯 PERÍODO: Enero 2024 → Septiembre 2025")
        
        if not CLIENT_ID or not SECRET_ID:
            print("❌ Credenciales no configuradas en .env")
            return
        
        fecha_hasta = datetime.now().strftime("%Y-%m-%d")
        print(f"🔧 Configuración:")
        print(f"   • Desde: {CONFIG_FECHAS['FECHA_DESDE']}")
        print(f"   • Hasta: {fecha_hasta}")
        print(f"   • Cliente: {CLIENT_ID[:15]}...")
        
        
        token = get_token()
        print("✅ Token obtenido\n")
        
        datos = {}
        
        print("📥 DESCARGA CON MÉTODOS OPTIMIZADOS:")
        print("=" * 45)
        
        # Clasificar endpoints por método
        endpoints_con_fechas = ['factura_compra', 'factura_venta', 'cobros', 'retenciones']
        endpoints_sin_fechas = ['clientes', 'cuentas', 'categorias_cuentas', 'pagos']
        endpoints_especiales = ['asiento_contable']  
        
        total_endpoints = len(ENDPOINTS_FUNCIONALES)
        current = 0
        
        MODO_COMPLETO_ASIENTOS = True  
        
        for nombre, endpoint in ENDPOINTS_FUNCIONALES.items():
            current += 1
            print(f"[{current}/{total_endpoints}] 🎯 {nombre.upper()}")
            
            if nombre in endpoints_con_fechas:
                
                datos[nombre] = get_data_monthly_chunks_only(token, nombre, endpoint)
                
            elif nombre in endpoints_especiales:
                
                if MODO_COMPLETO_ASIENTOS:
                    
                    datos[nombre] = get_asientos_contables_con_detalle_mejorado(token, nombre, endpoint)
                else:
                    
                    datos[nombre] = get_asientos_contables_debug_solo(token, nombre, endpoint)
                    
            else:
                
                datos[nombre] = get_data_simple_for_catalogs(token, nombre, endpoint)
            
            
            time.sleep(1)
        
        
        print("=" * 60)
        print("📊 RESUMEN FINAL")
        print("=" * 60)
        
        total_registros = 0
        endpoints_exitosos = 0
        
        for nombre, data in datos.items():
            count = len(data) if data else 0
            total_registros += count
            if count > 0:
                endpoints_exitosos += 1
            
            status = "✅" if count > 0 else "❌"
            
            
            if nombre in endpoints_especiales:
                metodo = "(diagnóstico automático)" if MODO_COMPLETO_ASIENTOS else "(debug estructura)"
            elif nombre in endpoints_con_fechas:
                metodo = "(chunks mensuales)"
            else:
                metodo = "(catálogo simple)"
                
            print(f"{status} {nombre:<20}: {count:>6,} registros {metodo}")
        
        print(f"\n🎯 Endpoints exitosos: {endpoints_exitosos}/{len(datos)}")
        print(f"📈 Total registros: {total_registros:,}")
        
        if total_registros > 0:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"xubio_diagnostico_{timestamp}.xlsx"
            reporte_filename = f"reporte_diagnostico_{timestamp}.txt"
            
            print(f"\n💾 EXPORTANDO...")
            exportar_a_excel_simple(datos, excel_filename)
            generar_reporte_mensual(datos, reporte_filename)
            
            print(f"\n🎉 ¡COMPLETADO!")
            print(f"📊 Excel: {excel_filename}")
            print(f"📄 Reporte: {reporte_filename}")
            print(f"⚡ Método: Chunks mensuales + diagnóstico automático")
            print(f"⏱️ Período: Enero 2024 → {fecha_hasta}")
            print(f"🔍 Asientos contables: {'Diagnóstico completo' if MODO_COMPLETO_ASIENTOS else 'Solo estructura'}")
            
        else:
            print(f"\n⚠️ No se obtuvieron datos")
            
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()